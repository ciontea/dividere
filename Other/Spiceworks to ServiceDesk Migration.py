#Purpose: You have an SQLite database you need to export to Excel .xlsx format since .csv may not always be reliable due to data stored
#Pre req: pip install pandas #THIS MUST BE RUN IN REGULAR POWERSHELL, NOT PYTHON
#Pre req: pip install openpyxl #THIS MUST BE RUN IN REGULAR POWERSHELL, NOT PYTHON
#Pre req: pip install xlwt #THIS MUST BE RUN IN REGULAR POWERSHELL, NOT PYTHON
#Note: With Python, if, for and other loop statements must not have spaces between lines of code
#Note: In Python, indentation is super important and can cause code not to work

#######################
#VARIABLES
#######################
databasePath = r"C:\Users\dciontea\Downloads\db\spiceworks_prod.db" #The "r" in front of the path is needed or Python script will break since it can't handle back slash characters
exportPath = r"C:\Users\dciontea\Downloads\output.xlsx" #The "r" in front of the path is needed or Python script will break since it can't handle back slash characters

#######################
#EXPORTING THE TICKETS FROM THE DATABASE WITH COMMENTS AND ATTACHMENT NAMES
#######################

import sqlite3
import pandas as pd
import openpyxl
import xlrd #Allows export to .xlsx
import xlwt #Allows export to .xls (This is being deprecated)

# Connect to the database file
conn = sqlite3.connect(databasePath)

# Get a cursor to execute queries on the database
cursor = conn.cursor()

# Execute a multi-line SQL query
query = """
SELECT
tickets.id,
tickets.summary AS subject,
--Adding new column for ticket priority levels so it maps properly to Servicedesk
    CASE
        WHEN tickets.priority = 1 THEN 'High'
        WHEN tickets.priority = 2 THEN 'Medium'
        WHEN tickets.priority = 3 THEN 'Low'
        ELSE tickets.priority
    END AS priority,
--Adding new column for ticket categories that no longer exist or should be blank like "Unspecified"
    CASE
        WHEN tickets.category = 'Unspecified' THEN ''
        --Non-existing categories that were used in tickets. Changing old category into new category after the THEN statement
        WHEN tickets.category = 'Access Request' THEN 'File Permissions'
        WHEN tickets.category = 'Adobe Acrobat Pro' THEN 'Adobe Acrobat'
        WHEN tickets.category = 'Cable Request' THEN 'Purchase Request'
        ELSE tickets.category
    END AS category,
--Adding new column for ticket status so it maps properly to Servicedesk. tickets.status is case sensitive!
    CASE
        WHEN tickets.c_status IN ('Closed - Resolved', 'Closed - After Hours', 'Closed - Acknowledged', 'Closed - Unresolved', 'Closed - Duplicate') THEN 'Closed'
        WHEN tickets.status = 'closed' AND tickets.c_status NOT IN ('Closed - Resolved', 'Closed - After Hours', 'Closed - Acknowledged', 'Closed - Unresolved', 'Closed - Duplicate') THEN 'Closed'
        WHEN tickets.status = 'closed' AND tickets.c_status IS NULL THEN 'Closed'
        WHEN tickets.status = 'open' AND tickets.c_status IS NULL THEN 'Pending'
        ELSE tickets.c_status
    END AS status,
--Grabbing the first name of who is assigned the ticket to complete it instead of their user ID number
    users1.first_name AS new_tickets_created_by_first_name,
--Grabbing the last name of who is assigned the ticket to complete it instead of their user ID number
    users1.last_name AS new_tickets_created_by_last_name,
--Grabbing the email of who is assigned the ticket to complete it instead of their user ID number
    users1.email AS new_tickets_created_by_email,
--Grabbing the first and last name of who is assigned the ticket to complete it instead of their user ID number. If first or last name is null, it will use their email address instead since that is always filled out
    (CASE
        WHEN users1.first_name IS NOT NULL AND users1.last_name IS NOT NULL THEN users1.first_name || ' ' || users1.last_name
        ELSE users1.email
    END) AS 'Requester Name',
--Grabbing the first name of who created the ticket to complete it instead of their user ID number
    users2.first_name AS new_tickets_assigned_to_first_name,
--Grabbing the last name of who created the ticket to complete it instead of their user ID number
    users2.last_name AS new_tickets_assigned_to_last_name,
--Grabbing the email of who created the ticket to complete it instead of their user ID number
    users2.email AS new_tickets_assigned_to_email,
--Grabbing the first and last name of who created the ticket instead of their user ID number. If first or last name is null, it will use their email address instead since that is always filled out
    (CASE
        WHEN users2.first_name IS NOT NULL AND users2.last_name IS NOT NULL THEN users2.first_name || ' ' || users2.last_name
        ELSE users2.email
    END) AS technician,
--Formatting Description Column better to be combined through Python
    ('DESCRIPTION: ' || tickets.description || CHAR(10)) AS new_tickets_description,
--Combining all ticket comments with who made the comment and the attachment file name if there is one and without the description at the beginning. Will be 2 columns. If first or last name is null, it will use their email address instead since that is always filled out
    GROUP_CONCAT(('COMMENT: Update From ' ||
        CASE
            WHEN users3.first_name IS NOT NULL AND users3.last_name IS NOT NULL THEN users3.first_name || ' ' || users3.last_name
            ELSE users3.email
        END || ' on ' || comments.created_at || ': ' || comments.body ||
        CASE
            WHEN comments.attachment_name IS NOT NULL THEN '------' || comments.attachment_name
            ELSE ''
        END), CHAR(10)) AS new_comments_combined, --CHAR(10) is a character constant that represents a line feed character (newline) in ASCII code
--Changing format of ticket created_at date to what Servicedesk Plus is expecting "dd MMM yyyy, hh:mm:ss". %m needs a little more work since there is no built in way to get Jun for month of June and only provides 06 by default
    strftime('%d ', tickets.created_at) || (substr('JanFebMarAprMayJunJulAugSepOctNovDec', 1 + 3*strftime('%m', tickets.created_at), -3) || strftime(' %Y, %H:%M:%S', tickets.created_at)) AS 'Created Date',
--Changing format of ticket closed_at date to what Servicedesk Plus is expecting "dd MMM yyyy, hh:mm:ss"
    strftime('%d ', tickets.closed_at) || (substr('JanFebMarAprMayJunJulAugSepOctNovDec', 1 + 3*strftime('%m', tickets.closed_at), -3) || strftime(' %Y, %H:%M:%S', tickets.closed_at)) AS 'Completed Date'

FROM tickets
JOIN comments ON tickets.id = comments.ticket_id
JOIN users AS users1 ON tickets.created_by = users1.id
JOIN users AS users2 ON tickets.assigned_to = users2.id
JOIN users AS users3 ON comments.created_by = users3.id
GROUP BY tickets.id --Removing duplicate rows of tickets.id since we want only one row for the ServiceDesk plus import
ORDER BY tickets.id ASC --Use DESC to get the most recent tickets (highest ticket number) on top
"""
cursor.execute(query) #Running the query above against the database
rows = cursor.fetchall() #Grabbing all rows from the database
conn.close() # Close the database connection


# Create a new .xlsx file
workbook = openpyxl.Workbook()

# Select the active worksheet
sheet = workbook.active

# Write the query result to the .xlsx file
for row in rows:
    sheet.append(row)

#######################
#MODIFYING THE OUTPUT.XLSX SO THAT THE HEADINGS ARE ADDED TO THE TOP
#######################

# Get the current number of rows in the sheet
row_count = sheet.max_row

# Insert a new row at the top of the sheet so the column headings can go above the data
sheet.insert_rows(1)

# Inserting the column headings
sheet["A1"] = "id"
sheet["B1"] = "subject"
sheet["C1"] = "priority"
sheet["D1"] = "category"
sheet["E1"] = "status"
sheet["F1"] = "new_tickets_created_by_first_name"
sheet["G1"] = "new_tickets_created_by_last_name"
sheet["H1"] = "new_tickets_created_by_email"
sheet["I1"] = "Requester Name"
sheet["J1"] = "new_tickets_assigned_to_first_name"
sheet["K1"] = "new_tickets_assigned_to_last_name"
sheet["L1"] = "new_tickets_assigned_to_email"
sheet["M1"] = "technician"
sheet["N1"] = "new_tickets_description"
sheet["O1"] = "new_comments_combined"
sheet["P1"] = "Created Date"
sheet["Q1"] = "Completed Date"
sheet["R1"] = "Description"

#######################
#MODIFYING THE OUTPUT.XLSX SO THAT THE COLUMNS new_tickets_description AND new_comments_combined ARE COMBINED INTO ONE CALLED new_tickets_comments_and_descriptions
#######################

# Combine the data in columns N "new_tickets_description" and O "new_comments_combined" into a new column R "new_tickets_description_and_comments"
column_new_tickets_description = int(14) #Column N
column_new_comments_combined = int(15) #Column O
column_new_tickets_description_and_comments = int(18) #Column R
for i in range(2, row_count + 1):
    sheet.cell(row=i, column=column_new_tickets_description_and_comments,
    value=str(sheet.cell(row=i, column=column_new_tickets_description).value) + " " + str(sheet.cell(row=i, column=column_new_comments_combined).value))

# Save the workbook
workbook.save(exportPath)

#######################
#MODIFYING THE OUTPUT.XLSX SO THAT IT CONTAINS FILLER TICKETS FOR THE ONES THAT HAVE BEEN DELETED USING PYTHON MODULES FOR EXCEL
#######################
#Important: Column names and almost everything is case sensitive

# Load the data into a pandas DataFrame
df = pd.read_excel(exportPath)

# Get the maximum ID number. Use df.columns to find all of the column names if you need to add more
max_id = df['id'].max()

# Create a list of all the ID numbers from 1 to max_id
id_list = [i for i in range(1, max_id + 1)]

# Filter the list to only include ID numbers that aren't already in the DataFrame
missing_ids = [id for id in id_list if id not in df['id'].values]

# Create a DataFrame for the missing IDs with "Blank Filler Ticket" in the Summary column
missing_data = {
    'id': missing_ids,
    'subject': ['Blank Filler Ticket'] * len(missing_ids),
    'priority': ['Medium'] * len(missing_ids),
    'category': ['Other'] * len(missing_ids),
    'status': ['Closed'] * len(missing_ids),
    'new_tickets_created_by_first_name': ['Backlog'] * len(missing_ids),
    'new_tickets_created_by_last_name': ['Tickets'] * len(missing_ids),
    'new_tickets_created_by_email': ['backlogtickets@reefersales.com'] * len(missing_ids),
    'Requester Name': ['Backlog Tickets'] * len(missing_ids),
    'new_tickets_assigned_to_first_name': ['Backlog'] * len(missing_ids),
    'new_tickets_assigned_to_last_name': ['Tickets'] * len(missing_ids),
    'new_tickets_assigned_to_email': ['backlogtickets@reefersales.com'] * len(missing_ids),
    'technician': ['Backlog Tickets'] * len(missing_ids),
    'new_tickets_description': ['Blank Filler Ticket'] * len(missing_ids),
    'new_comments_combined': ['Blank Filler Ticket'] * len(missing_ids),
    'Created Date': ['01 Jan 2022, 01:00:00'] * len(missing_ids),
    'Completed Date': ['01 Jan 2022, 01:05:00'] * len(missing_ids),
    'Description': ['Blank Filler Ticket'] * len(missing_ids)
}
missing_df = pd.DataFrame(missing_data)

# Concatenate the two DataFrames
df = pd.concat([df, missing_df], ignore_index=True)

# Sort the DataFrame by ID
df.sort_values(by='id', inplace=True)

# Save the modified DataFrame to a new XLSX file
df.to_excel(exportPath, index=False)

#######################
#SPLITTING THE FILE INTO MULTIPLE SMALLER
#######################

fileFormat = input("Would you like to split the large output file into smaller files. (y,n): ")
if fileFormat == "y":
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(exportPath)
    # Calculate the size of each row in bytes. Size does not reflect what Windows reports since this is the size in memory and is usally different
    row_size = df.memory_usage(index=True, deep=True).sum() / df.shape[0]
    # Set the desired file size in bytes
    fileSize = int(input("Size in MB you would like the largest file to be when splitting (Example Answer: 10). If you are importing these fiels to Servicedesk Plus, please make sure to type 10: "))
    file_size = fileSize * 1024 * 1024
    # Initialize the start and end indices for the first chunk
    start = 0
    end = 0
    # Initialize the cumulative size of each chunk
    cumulative_size = 0
    # Split the DataFrame into multiple DataFrames based on the desired file size
    df_list = []
    for i in range(df.shape[0]):
        cumulative_size += row_size
        end = i
        if cumulative_size >= file_size or i == df.shape[0] - 1:
            df_list.append(df.iloc[start:end+1, :])
            start = end + 1
            cumulative_size = 0
    fileFormat = input("Which file format would you like to export to, recommended is .xls for export to Servicedesk. (xls,xlsx): ")
    if fileFormat == "xlsx":
        # Save each DataFrame to a separate Excel file (.xlsx)
        for i, df_chunk in enumerate(df_list):
            df_chunk.to_excel(fr"C:\Users\dciontea\Downloads\output_split_{i}.xlsx", index=False) #f in front of the path stands for file and is needed
    else:
        # Save each DataFrame to a separate Excel file (.xls)
        for i, df_chunk in enumerate(df_list):
            writer = pd.ExcelWriter(fr"C:\Users\dciontea\Downloads\output_split_{i}.xls", engine='xlwt')
            df_chunk.to_excel(writer, index=False)
            writer.save()
